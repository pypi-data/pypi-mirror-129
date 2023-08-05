from io import BytesIO
from typing import Optional, List, Union

from openpyxl import \
    Workbook, \
    load_workbook

from xlsx_lib.domain.power_supply.power_supply_component import PowerSupplyComponent
from xlsx_lib.domain.power_supply.power_supply_sheet import PowerSupplySheet
from xlsx_lib.domain.frame.frame_element import FrameElement
from xlsx_lib.domain.frame.frame_sheet import FrameSheet
from xlsx_lib.domain.motorcycle_model.motorcycle_model import MotorcycleModel
from xlsx_lib.domain.motorcycle_model.sheetnames import Sheetnames

from xlsx_lib.domain.engine.engine_section import EngineSection
from xlsx_lib.domain.engine.engine_sheet import EngineSheet

from xlsx_lib.domain.electronic.electronic_element import ElectronicElement
from xlsx_lib.domain.electronic.electronic_sheet import ElectronicSheet

from xlsx_lib.domain.generic_replacements.generic_replacement_sheet import GenericReplacementsSheet
from xlsx_lib.domain.generic_replacements.replacement import Replacement

from xlsx_lib.domain.tightening_specifications.tightening_specifications_sheet import TighteningSpecificationsSheet
from xlsx_lib.domain.tightening_specifications.specification_element import SpecificationElement

sheetnames_relationships = {
    "MOT": Sheetnames.ENGINE,
    "REC. GENERICOS": Sheetnames.GENERIC_REPLACEMENTS,
    "ELEC": Sheetnames.ELECTRONIC,
    "PARES APRIETE": Sheetnames.TIGHTENING_TORQUES,
    "CHAS": Sheetnames.FRAME,
    "ALIM": Sheetnames.POWER_SUPPLY,
}


class MotorcycleModelWorkbook:
    motorcycle_model: Optional[MotorcycleModel]

    def __init__(
            self,
            file: Union[BytesIO, str],
            filename: Optional[str],
    ):
        workbook: Workbook = load_workbook(filename=file)

        model_name: str = filename[filename.rfind("/") + 1:filename.rfind(".")] \
            .replace("FICHA ", "") \
            .strip() \
            .replace("  ", " ")

        generic_replacements: Optional[List[Replacement]] = None
        electronic_elements: Optional[List[ElectronicElement]] = None
        tightening_specifications: Optional[List[SpecificationElement]] = None
        engine_sections: Optional[List[EngineSection]] = None
        frame_elements: Optional[List[FrameElement]] = None
        power_supply_components: Optional[List[PowerSupplyComponent]] = None

        for sheetname in [key for key in workbook.sheetnames
                          if key in sheetnames_relationships]:
            if Sheetnames.ENGINE == sheetnames_relationships[sheetname]:
                engine_sheet: EngineSheet = EngineSheet(worksheet=workbook[sheetname])
                engine_sections: List[EngineSection] = engine_sheet.engine_sections
                # TODO: Check if EngineSheet have distribution

            elif Sheetnames.GENERIC_REPLACEMENTS == sheetnames_relationships[sheetname]:
                generic_replacements = \
                    GenericReplacementsSheet(worksheet=workbook[sheetname]).get_generic_replacements()

            elif Sheetnames.ELECTRONIC == sheetnames_relationships[sheetname]:
                electronic_elements = \
                    ElectronicSheet(worksheet=workbook[sheetname]).get_electronic_elements()

            elif Sheetnames.TIGHTENING_TORQUES == sheetnames_relationships[sheetname]:
                tightening_specifications = \
                    TighteningSpecificationsSheet(worksheet=workbook[sheetname]).get_specification_elements()

            elif Sheetnames.FRAME == sheetnames_relationships[sheetname]:
                frame_elements = FrameSheet(worksheet=workbook[sheetname]).frame_elements

            elif Sheetnames.POWER_SUPPLY == sheetnames_relationships[sheetname]:
                power_supply_components = PowerSupplySheet(worksheet=workbook[sheetname]).components

        self.motorcycle_model = MotorcycleModel(
            model_name=model_name,
            generic_replacements=generic_replacements,
            tightening_specifications=tightening_specifications,
            electronic=electronic_elements,
            engine=engine_sections,
            frame=frame_elements,
            power_supply=power_supply_components,
        )

