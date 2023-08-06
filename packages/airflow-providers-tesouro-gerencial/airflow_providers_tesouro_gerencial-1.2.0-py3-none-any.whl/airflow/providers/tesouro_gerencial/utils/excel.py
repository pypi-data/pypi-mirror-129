from copy import copy

import openpyxl


def consolidar_tabela(livro_excel: openpyxl.Workbook) -> openpyxl.Workbook:
    '''Consolida tabela de planilha do Excel vinda do Tesouro Gerencial.

    O arquivo gerado pelo Tesouro Gerencial vem com título e outros metadados
    acima da tabela. Esta funcão tem por fim:

    i. Transformar células mescladas em células com o valor repetido
    ii. Isolar tabela na planilha, para fácil leitura com biblioteca `pandas`;
    iii. Adicionar coluna com metadados
    '''
    novo_livro_excel = copy(livro_excel)
    planilha = novo_livro_excel.worksheets[0]

    # Resolve células mescladas ao preencher seus valores
    grupos = planilha.merged_cells.ranges.copy()

    for grupo in grupos:
        esquerda, cima, direita, baixo = grupo.bounds
        planilha.unmerge_cells(grupo.coord)
        referencia = planilha.cell(cima, esquerda)

        for linha in planilha.iter_rows(
            min_col=esquerda, min_row=cima, max_col=direita, max_row=baixo
        ):
            for celula in linha:
                celula.value = referencia.value
                if referencia.has_style:
                    celula._style = copy(referencia._style)

    # Separa metadados de dados (antes de última linha vazia e depois de
    # última linha vazia)
    linhas_vazias = [
        n_linha + 1
        for n_linha, linha in enumerate(planilha.iter_rows())
        if all(celula.value is None for celula in linha)
    ]

    if linhas_vazias:
        divisor = linhas_vazias[-1]

        metadados = '\n'.join(filter(None, (
            planilha.cell(linha, 1).value
            for linha in range(1, divisor)
        )))

        planilha.delete_rows(1, divisor)
    else:
        metadados = None

    # Resolve cabeçalho multi-linha
    if planilha.cell(1, 1).fill.start_color.index \
            == planilha.cell(2, 1).fill.start_color.index:
        for n_coluna in range(1, planilha.max_column + 1):
            cima = planilha.cell(1, n_coluna)
            baixo = planilha.cell(2, n_coluna)

            if cima.value != baixo.value:
                baixo.value = cima.value + ' - ' + baixo.value

        planilha.delete_rows(1, 1)

    # Adiciona nova coluna com metadados
    nova_coluna = planilha.max_column + 1
    planilha.cell(1, nova_coluna, 'Metadados')

    for linha in range(2, planilha.max_row + 1):
        planilha.cell(linha, nova_coluna, metadados)

    return novo_livro_excel
