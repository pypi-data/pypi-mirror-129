import re
import json
import pytest
import logging
import requests
from pathlib import Path
from openpyxl import load_workbook
from caterpillar_log import logger

log = logging.getLogger("caterpillar_log")


class ApiTest(object):
    """
    基于pytest的数据驱动的接口自动化框架
    数据驱动：
        文件类型：         .xlsx
        xlsx文件头字段：    case_id       case_name       url                  method      headers     data        dependency     assert      setup                        teardown                   mark
        xlsx文件头含义      用例id         用例名称        接口url(不含ip:port)    方法        头部信息      body体      数据依赖         断言        接口执行之前函数自定义函数名      接口执行之后的自定义函数名      标记

    存储接口数据的结构体，case_id需要保持唯一性，实例如下：
        case_data={
            "case_id_1":{
                "headers":{},
                "data":{},
                "response":{},
                "status_code":200
            },
            "case_id_2":{
                "headers":{},
                "data":{},
                "response":{},
                "status_code":200
            }
        }
    """
    case_data = {}

    @staticmethod
    def get_all_xlsx_files(current_file, current_test_datas_dir):
        xlsx_files = []
        test_datas_dir = Path(current_file).resolve().parent / current_test_datas_dir
        for xlsx_file in test_datas_dir.iterdir():
            if str(xlsx_file).endswith(".xlsx"):
                xlsx_files.append(xlsx_file)
        return xlsx_files

    @staticmethod
    def get_all_test_datas(current_file, current_test_datas_dir):
        test_datas = []
        for xlsx in ApiTest.get_all_xlsx_files(current_file, current_test_datas_dir):
            workbook = load_workbook(xlsx)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.rows:
                    if not row[0].value or "case_id" in row[0].value or len(row[0].value.strip()) == 0:
                        continue
                    test_datas.append([elem.value for elem in row])
        return test_datas

    @staticmethod
    def loads(data):
        try:
            data_new = json.loads(data)
        except Exception as e:
            log.warning(f"在执行json.loads(data)时报错，返回空字典，data为：{data}，异常信息为{str(e)}")
            data_new = {}
        return data_new

    @staticmethod
    def request(url="", headers=None, data=None, method="GET", json=None,**kwargs):
        method = method.strip().upper()
        response = None
        if method == "GET":
            if data:
                params = data
            else:
                params = {}
            response = requests.get(url, params=params, headers=headers)
            log.info(f"接口请求url：{url}，请求方法method：{method}，请求返回状态码：status_code：{response.status_code}")
        elif method == "POST":
            if json is not None:
                response = requests.post(url, json=json, headers=headers, **kwargs)
            else:
                response = requests.post(url, data=data, headers=headers, json=json,**kwargs)
            log.info(f"接口请求url：{url}，请求方法method：{method}，请求返回状态码：status_code：{response.status_code}")
        elif method == "PUT":
            if json is not None:
                response = requests.put(url, json=data, headers=headers)
            else:
                response = requests.put(url, data=data, headers=headers)
        elif method == "DELETE":
            response = requests.delete(url, data=data, headers=headers)
            log.info(f"接口请求url：{url}，请求方法method：{method}，请求返回状态码：status_code：{response.status_code}")
        else:
            log.warning(f"目前请求方法支持post,get,put,delete，不区分大小写，当前请求方法不在上述之列，请求方法未：{method}")
        log.info(
            f"url:{url}, headers:{headers}, data:{data}, method:{method}, status_code:{response.status_code}, response.text:{response.text}")
        return response

    @staticmethod
    def save_data_to_case_data(case_id, data):
        if case_id not in ApiTest.case_data.keys():
            ApiTest.case_data[case_id] = {}
        ApiTest.case_data[case_id]["data"] = data

    @staticmethod
    def save_response_to_case_data(case_id, response):
        if case_id not in ApiTest.case_data.keys():
            ApiTest.case_data[case_id] = {}
        ApiTest.case_data[case_id]["headers"] = response.headers
        try:
            ApiTest.case_data[case_id]["response"] = response.json()
        except Exception as e:
            log.warning(f"读取接口垫额返回的json数据异常，这里默认失败时返回空字典，异常信息如下：{str(e)}")
            ApiTest.case_data[case_id]["response"] = {}
        ApiTest.case_data[case_id]["response"]["text"] = response.text
        ApiTest.case_data[case_id]["status_code"] = response.status_code

    @staticmethod
    def get_value_by_dot_expr(dot_expr):
        if dot_expr.startswith('"') and dot_expr.endswith('"'):
            return dot_expr.strip('"')
        elif dot_expr.startswith("'") and dot_expr.endswith("'"):
            return dot_expr.strip("'")
        elif "." in dot_expr:
            value=ApiTest.case_data
            for key in dot_expr.split("."):
                key = key.strip()
                try:
                    temp = re.search("(.*?)\\[(\\d+)\\]$", key)
                    if temp:
                        key = temp.group(1)
                        index = int(temp.group(2))
                        value = value[key][index]
                    else:
                        value = value[key]
                except Exception as e:
                    log.exception(f"当前value的值为：{value},key的值为：{key}，异常信息为：{str(e)}")
                    raise e
            return value
        else:
            return dot_expr

    @staticmethod
    def deal_dependency(headers, data, dependencies):
        if not dependencies:
            log.info("当前接口对其他接口没有依赖")
            return headers, data
        dependencies = dependencies.strip()
        for dependency in dependencies.split(";"):
            dependency = dependency.strip()
            if "=" not in dependency.strip():
                log.warning(f"接口对其他接口的依赖条件需要使用等号=赋值，当前依赖中未检测到=，所以这里将忽略此依赖，当前依赖信息为：{dependency}")
                continue
            src, dest = dependency.split("=")
            src = src.strip()
            dest = dest.strip()
            if "," in dest:
                dest_datas=[]
                for sub_dest in dest.split(","):
                    sub_value=ApiTest.get_value_by_dot_expr(sub_dest)
                    dest_datas.append(sub_value)
                value=",".join(dest_datas)
            else:
                value=ApiTest.get_value_by_dot_expr(dest)
            # 获取dest指定的值

            # 将获取到的value赋值
            dependency_list = src.split(".")
            if dependency_list[0] == "headers":
                obj = headers
                for key in dependency_list[1:-1]:
                    if key not in obj:
                        obj[key] = {}
                    obj = obj[key]
                obj[dependency_list[-1]] = value

            if dependency_list[0] == "data":
                obj = data
                for key in dependency_list[1:-1]:
                    temp = re.search("(.*?)\\[(\\d+)\\]$", key)
                    if temp:
                        key = temp.group(1)
                        index = int(temp.group(2))
                        if key not in obj.keys():
                            obj[key] = []
                            for i in range(index + 1):
                                obj[key].append({})
                        else:
                            if len(obj[key]) < index + 1:
                                for i in range(len(obj[key]), index + 1):
                                    obj[key].append({})
                        obj = obj[key][index]
                    else:
                        if key not in obj:
                            obj[key] = {}
                        obj = obj[key]
                last_key = dependency_list[-1]
                temp = re.search("(.*?)\\[(\\d+)\\]$", last_key)
                if temp:
                    key = temp.group(1)
                    index = int(temp.group(2))
                    if key not in obj.keys():
                        obj[key] = []
                        for i in range(index + 1):
                            obj[key].append({})
                    else:
                        if len(obj[key]) < index + 1:
                            for i in range(len(obj[key]), index + 1):
                                obj[key].append({})
                    obj[key][index] = value
                else:
                    if last_key not in obj:
                        obj[last_key] = {}
                    obj[last_key] = value
        return headers, data

    @staticmethod
    def api_assert(assert_str):
        """
        xlsx 文件中的断言列，为字符串，直接转换为可以进行assert断言的语句
        :param assert_str: 字符串
        :return:
        """
        assert_str = assert_str.strip()

        for assert_seg in assert_str.split(";"):
            assert_seg = assert_seg.strip()
            op = ApiTest.get_op(assert_seg)
            if not op:
                log.warning("因未识别出断言语句中的断言运算操作符，所以这里默认断言为TRUE")
                assert True
                return
            left_path, right_path = assert_seg.split(op)
            left_path = left_path.strip()
            right_path = right_path.strip()
            if left_path.startswith('"') and left_path.endswith('"'):
                left_value = left_path.strip('"')
            elif left_path.startswith("'") and left_path.endswith("'"):
                left_value = left_path.strip("'")
            elif "." in left_path:
                left_value = ApiTest.case_data
                for key in left_path.split("."):
                    try:
                        temp = re.search("(.*?)\\[(\\d+)\\]$", key)
                        if temp:
                            key = temp.group(1)
                            index = int(temp.group(2))
                            left_value = left_value[key][index]
                        else:
                            left_value = left_value[key]
                    except Exception as e:
                        log.exception(f"left_value的值为:{left_value},key的值为：{key}，异常信息为：{str(e)}")
                        raise e
            else:
                left_value = left_path

            if right_path.startswith('"') and right_path.endswith('"'):
                right_value = right_path.strip('"')
            elif right_path.startswith("'") and right_path.endswith("'"):
                right_value = right_path.strip("'")
            elif "." in right_path:
                right_value = ApiTest.case_data
                for key in right_path.split("."):
                    try:
                        temp = re.search("(.*?)\\[(\\d+)\\]$", key)
                        if temp:
                            key = temp.group(1)
                            index = int(temp.group(2))
                            right_value = right_value[key][index]
                        else:
                            right_value = right_value[key]
                    except Exception as e:
                        log.exception(f"right_value的值为：{right_value},key的值为：{key}，异常信息为：{str(e)}")
                        raise e
            else:
                right_value = right_path

            ApiTest.api_assert_op(left_value, op, right_value)

    @staticmethod
    def get_op(assert_str):
        if "==" in assert_str:
            return "=="
        elif ">=" in assert_str:
            return ">="
        elif "<=" in assert_str:
            return "<="
        elif "!=" in assert_str:
            return "!="
        elif ">" in assert_str:
            return ">"
        elif "<" in assert_str:
            return "<"
        elif " not in " in assert_str:
            return " not in "
        elif " in " in assert_str:
            return " in "
        elif " is not " in assert_str:
            return " is not "
        elif " is " in assert_str:
            return " is "
        else:
            log.warning("未解析出断言语句中的断言运算操作符，当前支持>=,<=,==,!=,not in,in,is not,is 运算符")
            return None

    @staticmethod
    def api_assert_op(left_value, op, right_value):
        if op == "==":
            if isinstance(left_value, int) or isinstance(right_value, int):
                left_value = int(left_value)
                right_value = int(right_value)
            assert left_value == right_value
        elif op == ">=":
            if isinstance(left_value, int) or isinstance(right_value, int):
                left_value = int(left_value)
                right_value = int(right_value)
            assert left_value >= right_value
        elif op == "<=":
            if isinstance(left_value, int) or isinstance(right_value, int):
                left_value = int(left_value)
                right_value = int(right_value)
            assert left_value <= right_value
        elif op == ">":
            if isinstance(left_value, int) or isinstance(right_value, int):
                left_value = int(left_value)
                right_value = int(right_value)
            assert left_value > right_value
        elif op == "<":
            if isinstance(left_value, int) or isinstance(right_value, int):
                left_value = int(left_value)
                right_value = int(right_value)
            assert left_value < right_value
        elif op == "!=":
            if isinstance(left_value, int) or isinstance(right_value, int):
                left_value = int(left_value)
                right_value = int(right_value)
            assert left_value != right_value
        elif op == " not in ":
            if isinstance(left_value, str) or isinstance(right_value, str):
                if isinstance(left_value, bool):
                    left_value = (str(left_value)).lower()
                else:
                    left_value = str(left_value)
                if isinstance(right_value, bool):
                    right_value = (str(right_value)).lower()
                else:
                    right_value = str(right_value)
            assert left_value not in right_value
        elif op == " in ":
            if isinstance(left_value, str) or isinstance(right_value, str):
                if isinstance(left_value,bool):
                    left_value = (str(left_value)).lower()
                else:
                    left_value = str(left_value)
                if isinstance(right_value,bool):
                    right_value = (str(right_value)).lower()
                else:
                    right_value = str(right_value)
            assert left_value in right_value
        elif op == " is not ":
            assert left_value is not right_value
        elif op == " is ":
            assert left_value is right_value
        else:
            log.warning("未解析出断言语句中的断言运算操作符，当前支持>=,<=,==,!=,not in,in,is not,is 运算符")
            assert True

    @staticmethod
    def run(case_id, case_name, url, method, headers, data, dependency, assert_str, mark):
        if "norun" in mark or "skip" in mark:
            pytest.skip("mark列有norun或skip字样，跳过不执行")
            return
        headers = ApiTest.loads(headers)
        data = ApiTest.loads(data)
        headers, data = ApiTest.deal_dependency(headers, data, dependency)
        ApiTest.save_data_to_case_data(case_id, data)
        response = ApiTest.request(url, headers, data, method)
        ApiTest.save_response_to_case_data(case_id, response)
        ApiTest.api_assert(assert_str)
